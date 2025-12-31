#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Dashboard.xlsx 自動生成腳本
使用 openpyxl 建立完整的專案管理 Dashboard
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, LineChart, DoughnutChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime, timedelta
import openpyxl

def create_dashboard():
    """建立 Dashboard.xlsx"""
    wb = Workbook()
    
    # 移除預設 Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 建立所有 Sheets
    create_project_overview(wb)
    create_task_list(wb)
    create_gantt_chart(wb)
    create_weekly_report(wb)
    create_analytics(wb)
    
    # 儲存檔案
    wb.save('Dashboard.xlsx')
    print("✅ Dashboard.xlsx 已成功建立！")

def create_project_overview(wb):
    """Sheet 1: 專案總覽"""
    ws = wb.create_sheet("專案總覽", 0)
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    
    # 標題樣式
    title_font = Font(name='微軟正黑體', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    
    # 基本資訊區
    ws['A1'] = '專案名稱'
    ws['B1'] = 'I&C Project Management'
    ws['A2'] = '專案經理'
    ws['B2'] = 'benchen1981@smail.nchu.edu.tw'
    ws['A3'] = '開始日期'
    ws['B3'] = datetime(2025, 1, 1)
    ws['B3'].number_format = 'yyyy-mm-dd'
    ws['A4'] = '預計結束日期'
    ws['B4'] = datetime(2025, 6, 30)
    ws['B4'].number_format = 'yyyy-mm-dd'
    ws['A5'] = '實際結束日期'
    ws['B5'] = ''
    ws['A6'] = '專案狀態'
    ws['B6'] = '進行中'
    ws['A7'] = '整體進度'
    ws['B7'] = '=IFERROR(COUNTIF(任務清單!F:F,"已完成")/COUNTA(任務清單!F2:F100)*100,0)&"%"'
    
    # 套用樣式到標題
    for row in range(1, 8):
        ws[f'A{row}'].font = Font(name='微軟正黑體', bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # KPI 卡片區
    kpi_font = Font(name='微軟正黑體', size=36, bold=True)
    kpi_label_font = Font(name='微軟正黑體', size=12, bold=True)
    
    # 卡片 1: 總任務數
    ws['A9'] = '總任務數'
    ws['A9'].font = kpi_label_font
    ws['A10'] = '=COUNTA(任務清單!B2:B100)'
    ws['A10'].font = Font(name='微軟正黑體', size=36, bold=True, color='0066CC')
    ws.merge_cells('A9:B9')
    ws.merge_cells('A10:B11')
    
    # 卡片 2: 已完成
    ws['C9'] = '已完成'
    ws['C9'].font = kpi_label_font
    ws['C10'] = '=COUNTIF(任務清單!F:F,"已完成")'
    ws['C10'].font = Font(name='微軟正黑體', size=36, bold=True, color='00CC00')
    ws.merge_cells('C9:D9')
    ws.merge_cells('C10:D11')
    
    # 卡片 3: 進行中
    ws['A13'] = '進行中'
    ws['A13'].font = kpi_label_font
    ws['A14'] = '=COUNTIF(任務清單!F:F,"進行中")'
    ws['A14'].font = Font(name='微軟正黑體', size=36, bold=True, color='FF9900')
    ws.merge_cells('A13:B13')
    ws.merge_cells('A14:B15')
    
    # 卡片 4: 逾期任務
    ws['C13'] = '逾期任務'
    ws['C13'].font = kpi_label_font
    ws['C14'] = '=COUNTIFS(任務清單!E:E,"<"&TODAY(),任務清單!F:F,"<>已完成")'
    ws['C14'].font = Font(name='微軟正黑體', size=36, bold=True, color='FF0000')
    ws.merge_cells('C13:D13')
    ws.merge_cells('C14:D15')
    
    # 設定對齊
    for cell in ['A9', 'C9', 'A13', 'C13', 'A10', 'C10', 'A14', 'C14']:
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')

def create_task_list(wb):
    """Sheet 2: 任務清單"""
    ws = wb.create_sheet("任務清單")
    
    # 表頭
    headers = ['任務ID', '任務名稱', '負責人', '優先級', '截止日期', '狀態', '完成日期', '進度%', '類別', '備註']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='微軟正黑體', bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 設定欄寬
    column_widths = [10, 30, 12, 10, 12, 12, 12, 10, 10, 30]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 範例資料
    tasks = [
        ['001', '撰寫專案章程', '專案經理', '🔴高', datetime(2025, 1, 5), '已完成', datetime(2025, 1, 4), 100, '文件', '已完成並審核'],
        ['002', '建立WBS', '專案經理', '🔴高', datetime(2025, 1, 6), '進行中', '', 60, '管理', ''],
        ['003', '利害關係人訪談', '張三', '🟡中', datetime(2025, 1, 8), '待辦', '', 0, '管理', ''],
        ['004', '任務優先級排序', '李四', '🔴高', datetime(2025, 1, 10), '待辦', '', 0, '管理', ''],
        ['005', '設定專案里程碑', '專案經理', '🔴高', datetime(2025, 1, 12), '待辦', '', 0, '管理', ''],
        ['006', '每日站立會議', '專案經理', '🔴高', datetime(2025, 12, 31), '進行中', '', 50, '管理', '每日重複'],
        ['007', '週進度檢討會議', '專案經理', '🔴高', datetime(2025, 12, 31), '進行中', '', 50, '管理', '每週重複'],
        ['008', '技術架構設計', '王五', '🔴高', datetime(2025, 1, 15), '待辦', '', 0, '技術', ''],
        ['009', '資料庫設計', '趙六', '🟡中', datetime(2025, 1, 20), '待辦', '', 0, '技術', ''],
        ['010', '前端介面設計', '張三', '🟡中', datetime(2025, 1, 25), '待辦', '', 0, '技術', ''],
    ]
    
    for row_idx, task in enumerate(tasks, start=2):
        for col_idx, value in enumerate(task, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # 日期格式
            if col_idx in [5, 7] and isinstance(value, datetime):
                cell.number_format = 'yyyy-mm-dd'
            # 進度百分比格式
            if col_idx == 8:
                cell.number_format = '0"%"'
    
    # 條件格式 - 優先級
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    
    ws.conditional_formatting.add('D2:D100',
        CellIsRule(operator='equal', formula=['"🔴高"'], fill=red_fill, font=Font(color='FFFFFF')))
    ws.conditional_formatting.add('D2:D100',
        CellIsRule(operator='equal', formula=['"🟡中"'], fill=yellow_fill))
    ws.conditional_formatting.add('D2:D100',
        CellIsRule(operator='equal', formula=['"🟢低"'], fill=green_fill))
    
    # 條件格式 - 狀態
    gray_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    blue_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    orange_fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
    green_fill2 = PatternFill(start_color='00CC00', end_color='00CC00', fill_type='solid')
    
    ws.conditional_formatting.add('F2:F100',
        CellIsRule(operator='equal', formula=['"待辦"'], fill=gray_fill))
    ws.conditional_formatting.add('F2:F100',
        CellIsRule(operator='equal', formula=['"進行中"'], fill=blue_fill, font=Font(color='FFFFFF')))
    ws.conditional_formatting.add('F2:F100',
        CellIsRule(operator='equal', formula=['"審核中"'], fill=orange_fill, font=Font(color='FFFFFF')))
    ws.conditional_formatting.add('F2:F100',
        CellIsRule(operator='equal', formula=['"已完成"'], fill=green_fill2, font=Font(color='FFFFFF')))
    
    # 條件格式 - 截止日期 (逾期)
    red_font = Font(color='FF0000', bold=True)
    ws.conditional_formatting.add('E2:E100',
        FormulaRule(formula=['AND($E2<TODAY(),$F2<>"已完成")'], font=red_font))
    
    # 條件格式 - 截止日期 (3天內)
    ws.conditional_formatting.add('E2:E100',
        FormulaRule(formula=['AND($E2<=TODAY()+3,$E2>=TODAY(),$F2<>"已完成")'], fill=yellow_fill))

def create_gantt_chart(wb):
    """Sheet 3: 甘特圖"""
    ws = wb.create_sheet("甘特圖")
    
    # 表頭
    headers = ['任務ID', '任務名稱', '負責人', '開始日期', '結束日期', '進度%']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='微軟正黑體', bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 日期欄位 (從 G1 開始，顯示 60 天)
    start_date = datetime(2025, 1, 1)
    for day in range(60):
        current_date = start_date + timedelta(days=day)
        cell = ws.cell(row=1, column=7+day, value=current_date)
        cell.number_format = 'mm/dd'
        cell.font = Font(name='微軟正黑體', size=8, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        ws.column_dimensions[get_column_letter(7+day)].width = 3
    
    # 設定前 6 欄欄寬
    column_widths = [10, 30, 12, 12, 12, 10]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 從任務清單引用資料
    for row in range(2, 12):
        ws[f'A{row}'] = f'=任務清單!A{row}'
        ws[f'B{row}'] = f'=任務清單!B{row}'
        ws[f'C{row}'] = f'=任務清單!C{row}'
        ws[f'D{row}'] = f'=任務清單!E{row}-7'  # 假設任務期間 7 天
        ws[f'E{row}'] = f'=任務清單!E{row}'
        ws[f'F{row}'] = f'=任務清單!H{row}'
        ws[f'F{row}'].number_format = '0"%"'
    
    # 條件格式 - 任務期間 (藍色)
    blue_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws.conditional_formatting.add('G2:BZ100',
        FormulaRule(formula=['AND($D2<=G$1,G$1<=$E2)'], fill=blue_fill))
    
    # 條件格式 - 已完成進度 (深藍色)
    dark_blue_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws.conditional_formatting.add('G2:BZ100',
        FormulaRule(formula=['AND($D2<=G$1,G$1<=$D2+($E2-$D2)*$F2/100)'], fill=dark_blue_fill))
    
    # 條件格式 - 今日線 (紅色)
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    ws.conditional_formatting.add('G2:BZ100',
        FormulaRule(formula=['G$1=TODAY()'], fill=red_fill))

def create_weekly_report(wb):
    """Sheet 4: 週報"""
    ws = wb.create_sheet("週報")
    
    # 週資訊區
    ws['A1'] = '週報期間'
    ws['B1'] = '=TEXT(TODAY()-WEEKDAY(TODAY())+1,"yyyy/mm/dd") & " - " & TEXT(TODAY()-WEEKDAY(TODAY())+7,"yyyy/mm/dd")'
    ws['A2'] = '報告日期'
    ws['B2'] = '=TODAY()'
    ws['B2'].number_format = 'yyyy-mm-dd'
    ws['A3'] = '報告人'
    ws['B3'] = '=專案總覽!B2'
    
    # 樣式
    for row in range(1, 4):
        ws[f'A{row}'].font = Font(name='微軟正黑體', bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # KPI 區
    ws['A5'] = '本週完成任務數'
    ws['B5'] = '=COUNTIFS(任務清單!G:G,">="&TODAY()-7,任務清單!G:G,"<="&TODAY(),任務清單!F:F,"已完成")'
    ws['A6'] = '本週新增任務數'
    ws['B6'] = '0'  # 需手動輸入
    ws['A7'] = '本週逾期任務數'
    ws['B7'] = '=COUNTIFS(任務清單!E:E,"<"&TODAY(),任務清單!F:F,"<>已完成")'
    ws['A8'] = '整體進度變化'
    ws['B8'] = '0%'  # 需手動輸入
    
    for row in range(5, 9):
        ws[f'A{row}'].font = Font(name='微軟正黑體', bold=True)
    
    # 本週重點
    ws['A10'] = '本週成就'
    ws['A10'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws['A10'].fill = PatternFill(start_color='00CC00', end_color='00CC00', fill_type='solid')
    ws.merge_cells('A10:J10')
    ws['A11'] = '[請輸入本週主要成就]'
    ws.merge_cells('A11:J13')
    
    ws['A15'] = '本週挑戰'
    ws['A15'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws['A15'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
    ws.merge_cells('A15:J15')
    ws['A16'] = '[請輸入本週遇到的挑戰]'
    ws.merge_cells('A16:J18')
    
    ws['A20'] = '下週計畫'
    ws['A20'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws['A20'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws.merge_cells('A20:J20')
    ws['A21'] = '[請輸入下週計畫]'
    ws.merge_cells('A21:J23')
    
    # 逾期任務清單
    ws['A25'] = '逾期任務清單'
    ws['A25'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws['A25'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    ws.merge_cells('A25:G25')
    
    headers = ['任務ID', '任務名稱', '負責人', '截止日期', '逾期天數', '狀態', '行動計畫']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=26, column=col, value=header)
        cell.font = Font(name='微軟正黑體', bold=True)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

def create_analytics(wb):
    """Sheet 5: 統計分析"""
    ws = wb.create_sheet("統計分析")
    
    # 任務狀態統計表
    ws['A1'] = '狀態'
    ws['B1'] = '數量'
    ws['A1'].font = Font(name='微軟正黑體', bold=True)
    ws['B1'].font = Font(name='微軟正黑體', bold=True)
    
    statuses = ['待辦', '進行中', '審核中', '已完成']
    for i, status in enumerate(statuses, start=2):
        ws[f'A{i}'] = status
        ws[f'B{i}'] = f'=COUNTIF(任務清單!F:F,"{status}")'
    
    # 優先級統計表
    ws['D1'] = '優先級'
    ws['E1'] = '數量'
    ws['D1'].font = Font(name='微軟正黑體', bold=True)
    ws['E1'].font = Font(name='微軟正黑體', bold=True)
    
    priorities = ['🔴高', '🟡中', '🟢低']
    for i, priority in enumerate(priorities, start=2):
        ws[f'D{i}'] = priority
        ws[f'E{i}'] = f'=COUNTIF(任務清單!D:D,"{priority}")'
    
    # 負責人工作量統計表
    ws['G1'] = '負責人'
    ws['H1'] = '任務數'
    ws['I1'] = '已完成數'
    for cell in ['G1', 'H1', 'I1']:
        ws[cell].font = Font(name='微軟正黑體', bold=True)
    
    assignees = ['張三', '李四', '王五', '趙六', '專案經理']
    for i, assignee in enumerate(assignees, start=2):
        ws[f'G{i}'] = assignee
        ws[f'H{i}'] = f'=COUNTIF(任務清單!C:C,"{assignee}")'
        ws[f'I{i}'] = f'=COUNTIFS(任務清單!C:C,"{assignee}",任務清單!F:F,"已完成")'
    
    # 類別統計表
    ws['K1'] = '類別'
    ws['L1'] = '數量'
    ws['K1'].font = Font(name='微軟正黑體', bold=True)
    ws['L1'].font = Font(name='微軟正黑體', bold=True)
    
    categories = ['管理', '技術', '文件']
    for i, category in enumerate(categories, start=2):
        ws[f'K{i}'] = category
        ws[f'L{i}'] = f'=COUNTIF(任務清單!I:I,"{category}")'
    
    # 建立圖表
    # 1. 任務狀態圓餅圖
    pie1 = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=5)
    data = Reference(ws, min_col=2, min_row=1, max_row=5)
    pie1.add_data(data, titles_from_data=True)
    pie1.set_categories(labels)
    pie1.title = "任務狀態分布"
    ws.add_chart(pie1, "A8")
    
    # 2. 優先級圓餅圖
    pie2 = PieChart()
    labels2 = Reference(ws, min_col=4, min_row=2, max_row=4)
    data2 = Reference(ws, min_col=5, min_row=1, max_row=4)
    pie2.add_data(data2, titles_from_data=True)
    pie2.set_categories(labels2)
    pie2.title = "優先級分布"
    ws.add_chart(pie2, "D8")
    
    # 3. 負責人工作量橫條圖
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "團隊成員工作量"
    bar.y_axis.title = '任務數'
    bar.x_axis.title = '負責人'
    
    data3 = Reference(ws, min_col=8, min_row=1, max_row=6)
    cats3 = Reference(ws, min_col=7, min_row=2, max_row=6)
    bar.add_data(data3, titles_from_data=True)
    bar.set_categories(cats3)
    ws.add_chart(bar, "G8")
    
    # 4. 類別分布圓餅圖
    pie3 = PieChart()
    labels3 = Reference(ws, min_col=11, min_row=2, max_row=4)
    data4 = Reference(ws, min_col=12, min_row=1, max_row=4)
    pie3.add_data(data4, titles_from_data=True)
    pie3.set_categories(labels3)
    pie3.title = "類別分布"
    ws.add_chart(pie3, "K8")

if __name__ == '__main__':
    print("🚀 開始建立 Dashboard.xlsx...")
    try:
        create_dashboard()
        print("\n📊 Dashboard.xlsx 建立完成！")
        print("\n包含以下 Sheets:")
        print("  1. 專案總覽 - KPI 卡片與基本資訊")
        print("  2. 任務清單 - 完整任務資料與條件格式")
        print("  3. 甘特圖 - 視覺化時間軸")
        print("  4. 週報 - 週報模板")
        print("  5. 統計分析 - 圖表與統計資料")
    except Exception as e:
        print(f"❌ 錯誤: {e}")
        import traceback
        traceback.print_exc()
