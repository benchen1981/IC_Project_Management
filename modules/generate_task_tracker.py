#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Task_Tracker.xlsx 自動生成腳本
簡化版任務追蹤表，用於日常快速更新
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime

def create_task_tracker():
    """建立 Task_Tracker.xlsx"""
    wb = Workbook()
    
    # 移除預設 Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 建立所有 Sheets
    create_my_tasks(wb)
    create_today_tasks(wb)
    create_this_week_tasks(wb)
    create_overdue_tasks(wb)
    
    # 儲存檔案
    wb.save('Task_Tracker.xlsx')
    print("✅ Task_Tracker.xlsx 已成功建立！")

def setup_task_sheet(ws, title, filter_formula_description):
    """設定任務工作表的共用格式"""
    # 標題
    ws['A1'] = title
    ws['A1'].font = Font(name='微軟正黑體', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30
    
    # 說明
    ws['A2'] = filter_formula_description
    ws['A2'].font = Font(name='微軟正黑體', size=10, italic=True)
    ws.merge_cells('A2:H2')
    
    # 表頭
    headers = ['任務ID', '任務名稱', '負責人', '優先級', '截止日期', '狀態', '進度%', '備註']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name='微軟正黑體', bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 設定欄寬
    column_widths = [10, 35, 12, 10, 12, 12, 10, 30]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 條件格式 - 優先級
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    
    ws.conditional_formatting.add('D4:D100',
        FormulaRule(formula=['$D4="🔴高"'], fill=red_fill, font=Font(color='FFFFFF')))
    ws.conditional_formatting.add('D4:D100',
        FormulaRule(formula=['$D4="🟡中"'], fill=yellow_fill))
    ws.conditional_formatting.add('D4:D100',
        FormulaRule(formula=['$D4="🟢低"'], fill=green_fill))
    
    # 條件格式 - 狀態
    gray_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    blue_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    orange_fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
    green_fill2 = PatternFill(start_color='00CC00', end_color='00CC00', fill_type='solid')
    
    ws.conditional_formatting.add('F4:F100',
        FormulaRule(formula=['$F4="待辦"'], fill=gray_fill))
    ws.conditional_formatting.add('F4:F100',
        FormulaRule(formula=['$F4="進行中"'], fill=blue_fill, font=Font(color='FFFFFF')))
    ws.conditional_formatting.add('F4:F100',
        FormulaRule(formula=['$F4="審核中"'], fill=orange_fill, font=Font(color='FFFFFF')))
    ws.conditional_formatting.add('F4:F100',
        FormulaRule(formula=['$F4="已完成"'], fill=green_fill2, font=Font(color='FFFFFF')))

def create_my_tasks(wb):
    """我的任務"""
    ws = wb.create_sheet("我的任務", 0)
    setup_task_sheet(ws, "📋 我的任務 (My Tasks)", "顯示所有指派給我的任務")
    
    # 範例資料 (實際使用時應從 Dashboard.xlsx 同步)
    tasks = [
        ['002', '建立WBS', '專案經理', '🔴高', datetime(2025, 1, 6), '進行中', '60%', ''],
        ['005', '設定專案里程碑', '專案經理', '🔴高', datetime(2025, 1, 12), '待辦', '0%', ''],
        ['006', '每日站立會議', '專案經理', '🔴高', datetime(2025, 12, 31), '進行中', '50%', '每日重複'],
        ['007', '週進度檢討會議', '專案經理', '🔴高', datetime(2025, 12, 31), '進行中', '50%', '每週重複'],
    ]
    
    for row_idx, task in enumerate(tasks, start=4):
        for col_idx, value in enumerate(task, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 5 and isinstance(value, datetime):
                cell.number_format = 'yyyy-mm-dd'

def create_today_tasks(wb):
    """今日任務"""
    ws = wb.create_sheet("今日任務")
    setup_task_sheet(ws, "📅 今日任務 (Today)", "顯示今日到期或需處理的任務")
    
    # 範例資料
    ws['A4'] = '006'
    ws['B4'] = '每日站立會議'
    ws['C4'] = '專案經理'
    ws['D4'] = '🔴高'
    ws['E4'] = datetime.now()
    ws['E4'].number_format = 'yyyy-mm-dd'
    ws['F4'] = '進行中'
    ws['G4'] = '50%'
    ws['H4'] = '每日 9:00 AM'

def create_this_week_tasks(wb):
    """本週任務"""
    ws = wb.create_sheet("本週任務")
    setup_task_sheet(ws, "📆 本週任務 (This Week)", "顯示本週到期的任務")
    
    # 範例資料
    tasks = [
        ['002', '建立WBS', '專案經理', '🔴高', datetime(2025, 1, 6), '進行中', '60%', ''],
        ['003', '利害關係人訪談', '張三', '🟡中', datetime(2025, 1, 8), '待辦', '0%', ''],
        ['004', '任務優先級排序', '李四', '🔴高', datetime(2025, 1, 10), '待辦', '0%', ''],
    ]
    
    for row_idx, task in enumerate(tasks, start=4):
        for col_idx, value in enumerate(task, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 5 and isinstance(value, datetime):
                cell.number_format = 'yyyy-mm-dd'

def create_overdue_tasks(wb):
    """逾期任務"""
    ws = wb.create_sheet("逾期任務")
    setup_task_sheet(ws, "🔔 逾期任務 (Overdue)", "顯示已逾期但未完成的任務")
    
    # 額外欄位：逾期天數
    ws['I3'] = '逾期天數'
    ws['I3'].font = Font(name='微軟正黑體', bold=True, color='FFFFFF')
    ws['I3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['I3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['I'].width = 12
    
    # 說明文字
    ws['A2'] = '⚠️ 這些任務需要立即處理！'
    ws['A2'].font = Font(name='微軟正黑體', size=10, bold=True, color='FF0000')

if __name__ == '__main__':
    print("🚀 開始建立 Task_Tracker.xlsx...")
    try:
        create_task_tracker()
        print("\n📋 Task_Tracker.xlsx 建立完成！")
        print("\n包含以下 Sheets:")
        print("  1. 我的任務 - 所有指派給我的任務")
        print("  2. 今日任務 - 今日需處理的任務")
        print("  3. 本週任務 - 本週到期的任務")
        print("  4. 逾期任務 - 已逾期未完成的任務")
    except Exception as e:
        print(f"❌ 錯誤: {e}")
        import traceback
        traceback.print_exc()
