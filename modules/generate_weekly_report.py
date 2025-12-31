#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Weekly_Report_Template.xlsx 自動生成腳本
週報模板，可自動填充本週日期範圍和統計數據
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime

def create_weekly_report_template():
    """建立 Weekly_Report_Template.xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = "週報"
    
    # 設定欄寬
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 標題區
    ws['A1'] = 'I&C Project Management'
    ws['A1'].font = Font(name='微軟正黑體', size=20, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 35
    
    ws['A2'] = '週進度報告 (Weekly Progress Report)'
    ws['A2'].font = Font(name='微軟正黑體', size=14, bold=True, color='FFFFFF')
    ws['A2'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A2:J2')
    ws.row_dimensions[2].height = 25
    
    # 週資訊區
    ws['A4'] = '週報期間'
    ws['B4'] = '=TEXT(TODAY()-WEEKDAY(TODAY())+1,"yyyy/mm/dd") & " - " & TEXT(TODAY()-WEEKDAY(TODAY())+7,"yyyy/mm/dd")'
    ws['A5'] = '報告日期'
    ws['B5'] = '=TODAY()'
    ws['B5'].number_format = 'yyyy-mm-dd'
    ws['A6'] = '報告人'
    ws['B6'] = 'benchen1981@smail.nchu.edu.tw'
    ws['A7'] = '專案狀態'
    ws['B7'] = '進行中'
    
    # 樣式
    for row in range(4, 8):
        ws[f'A{row}'].font = Font(name='微軟正黑體', bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        ws[f'B{row}'].alignment = Alignment(horizontal='left')
    
    # KPI 區標題
    ws['A9'] = '📊 本週關鍵指標 (Key Metrics)'
    ws['A9'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws['A9'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws['A9'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A9:J9')
    
    # KPI 卡片
    kpi_data = [
        ('本週完成任務數', '=COUNTIFS(Dashboard.xlsx!任務清單!G:G,">="&TODAY()-7,Dashboard.xlsx!任務清單!G:G,"<="&TODAY(),Dashboard.xlsx!任務清單!F:F,"已完成")', 'A11', 'B11', '00CC00'),
        ('本週新增任務數', '0', 'D11', 'E11', '0066CC'),
        ('本週逾期任務數', '=COUNTIFS(Dashboard.xlsx!任務清單!E:E,"<"&TODAY(),Dashboard.xlsx!任務清單!F:F,"<>已完成")', 'G11', 'H11', 'FF0000'),
        ('整體進度', '=IFERROR(COUNTIF(Dashboard.xlsx!任務清單!F:F,"已完成")/COUNTA(Dashboard.xlsx!任務清單!F2:F100)*100,0)&"%"', 'A13', 'B13', 'FF9900'),
        ('進行中任務', '=COUNTIF(Dashboard.xlsx!任務清單!F:F,"進行中")', 'D13', 'E13', '4472C4'),
        ('待辦任務', '=COUNTIF(Dashboard.xlsx!任務清單!F:F,"待辦")', 'G13', 'H13', 'CCCCCC'),
    ]
    
    for label, formula, label_cell, value_cell, color in kpi_data:
        ws[label_cell] = label
        ws[label_cell].font = Font(name='微軟正黑體', size=10, bold=True)
        ws[label_cell].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[value_cell] = formula
        ws[value_cell].font = Font(name='微軟正黑體', size=24, bold=True, color=color)
        ws[value_cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # 本週成就
    ws['A16'] = '✅ 本週成就 (Achievements)'
    ws['A16'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws['A16'].fill = PatternFill(start_color='00CC00', end_color='00CC00', fill_type='solid')
    ws['A16'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A16:J16')
    
    ws['A17'] = '• [請輸入本週主要成就 1]\n• [請輸入本週主要成就 2]\n• [請輸入本週主要成就 3]'
    ws['A17'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells('A17:J20')
    ws.row_dimensions[17].height = 60
    
    # 本週挑戰
    ws['A22'] = '⚠️ 本週挑戰 (Challenges)'
    ws['A22'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws['A22'].fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
    ws['A22'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A22:J22')
    
    ws['A23'] = '• [請輸入本週遇到的挑戰 1]\n• [請輸入本週遇到的挑戰 2]\n• [請輸入本週遇到的挑戰 3]'
    ws['A23'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells('A23:J26')
    ws.row_dimensions[23].height = 60
    
    # 下週計畫
    ws['A28'] = '🎯 下週計畫 (Next Week Plan)'
    ws['A28'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws['A28'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws['A28'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A28:J28')
    
    ws['A29'] = '• [請輸入下週計畫 1]\n• [請輸入下週計畫 2]\n• [請輸入下週計畫 3]'
    ws['A29'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells('A29:J32')
    ws.row_dimensions[29].height = 60
    
    # 逾期任務清單
    ws['A34'] = '🔴 逾期任務清單 (Overdue Tasks)'
    ws['A34'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws['A34'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    ws['A34'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A34:J34')
    
    # 逾期任務表頭
    headers = ['任務ID', '任務名稱', '負責人', '截止日期', '逾期天數', '狀態', '優先級', '行動計畫']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=35, column=col, value=header)
        cell.font = Font(name='微軟正黑體', bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 範例逾期任務 (實際使用時應自動從 Dashboard 篩選)
    ws['A36'] = '[自動從 Dashboard 篩選逾期任務]'
    ws.merge_cells('A36:H36')
    ws['A36'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A36'].font = Font(name='微軟正黑體', italic=True, color='999999')
    
    # 風險與議題
    ws['A38'] = '⚡ 風險與議題 (Risks & Issues)'
    ws['A38'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws['A38'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    ws['A38'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A38:J38')
    
    risk_headers = ['風險/議題', '影響程度', '發生機率', '應對措施', '負責人', '狀態']
    for col, header in enumerate(risk_headers, start=1):
        cell = ws.cell(row=39, column=col, value=header)
        cell.font = Font(name='微軟正黑體', bold=True)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 範例風險
    ws['A40'] = '[請輸入風險或議題]'
    ws['B40'] = '高/中/低'
    ws['C40'] = '高/中/低'
    ws['D40'] = '[應對措施]'
    ws['E40'] = '[負責人]'
    ws['F40'] = '進行中'
    
    # 團隊成員工作量
    ws['A43'] = '👥 團隊成員工作量 (Team Workload)'
    ws['A43'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws['A43'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    ws['A43'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A43:J43')
    
    workload_headers = ['成員', '總任務數', '已完成', '進行中', '待辦', '完成率', '工作負荷']
    for col, header in enumerate(workload_headers, start=1):
        cell = ws.cell(row=44, column=col, value=header)
        cell.font = Font(name='微軟正黑體', bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 範例成員資料
    members = ['張三', '李四', '王五', '趙六', '專案經理']
    for i, member in enumerate(members, start=45):
        ws[f'A{i}'] = member
        ws[f'B{i}'] = f'=COUNTIF(Dashboard.xlsx!任務清單!C:C,"{member}")'
        ws[f'C{i}'] = f'=COUNTIFS(Dashboard.xlsx!任務清單!C:C,"{member}",Dashboard.xlsx!任務清單!F:F,"已完成")'
        ws[f'D{i}'] = f'=COUNTIFS(Dashboard.xlsx!任務清單!C:C,"{member}",Dashboard.xlsx!任務清單!F:F,"進行中")'
        ws[f'E{i}'] = f'=COUNTIFS(Dashboard.xlsx!任務清單!C:C,"{member}",Dashboard.xlsx!任務清單!F:F,"待辦")'
        ws[f'F{i}'] = f'=IFERROR(C{i}/B{i}*100,0)&"%"'
        ws[f'G{i}'] = f'=IF(B{i}>10,"高",IF(B{i}>5,"中","低"))'
    
    # 頁尾
    ws['A51'] = '報告產生時間：=NOW()'
    ws['A51'].font = Font(name='微軟正黑體', size=9, italic=True, color='666666')
    ws.merge_cells('A51:J51')
    ws['A51'].alignment = Alignment(horizontal='center')
    
    ws['A52'] = '此報告由 I&C Project Management System 自動生成'
    ws['A52'].font = Font(name='微軟正黑體', size=9, italic=True, color='666666')
    ws.merge_cells('A52:J52')
    ws['A52'].alignment = Alignment(horizontal='center')
    
    # 儲存檔案
    wb.save('Weekly_Report_Template.xlsx')
    print("✅ Weekly_Report_Template.xlsx 已成功建立！")

if __name__ == '__main__':
    print("🚀 開始建立 Weekly_Report_Template.xlsx...")
    try:
        create_weekly_report_template()
        print("\n📄 Weekly_Report_Template.xlsx 建立完成！")
        print("\n包含以下區塊:")
        print("  • 週報期間與基本資訊")
        print("  • 本週關鍵指標 (KPI)")
        print("  • 本週成就")
        print("  • 本週挑戰")
        print("  • 下週計畫")
        print("  • 逾期任務清單")
        print("  • 風險與議題")
        print("  • 團隊成員工作量")
        print("\n💡 使用方式:")
        print("  1. 每週複製此模板建立新週報")
        print("  2. 自動填充的數據會從 Dashboard.xlsx 提取")
        print("  3. 手動填寫成就、挑戰、計畫等內容")
        print("  4. 匯出為 PDF 發送給團隊")
    except Exception as e:
        print(f"❌ 錯誤: {e}")
        import traceback
        traceback.print_exc()
